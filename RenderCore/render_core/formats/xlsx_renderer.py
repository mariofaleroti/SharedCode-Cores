from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..document import ReportDocument
from ..options import RenderOptions
from ..paths import ensure_parent
from ..result import RenderResult
from ..utils import default_output_path, render_value


DARK_FILL = "111827"
HEADER_FILL = "1F2937"
SECTION_FILL = "E5E7EB"
OK_FILL = "DCFCE7"
WARNING_FILL = "FEF3C7"
CRITICAL_FILL = "FEE2E2"
INFO_FILL = "DBEAFE"
BORDER_COLOR = "D1D5DB"


def render_xlsx(document: ReportDocument, options: RenderOptions) -> RenderResult:
    output_path = options.output_path or default_output_path(options.input_path, "xlsx", options.output_dir)
    output_path = output_path.expanduser().resolve()
    ensure_parent(output_path)

    workbook = Workbook()
    workbook.properties.title = document.title
    workbook.properties.subject = document.report_type
    workbook.properties.creator = "RenderCore"

    summary_sheet = workbook.active
    summary_sheet.title = "Resumen"
    _write_summary_sheet(summary_sheet, document)

    if document.report_type == "disk_smart":
        _write_disk_smart_overview(workbook, document)

    if document.diagnostics:
        _write_list_sheet(workbook.create_sheet("Diagnosticos"), document.diagnostics)

    if document.errors:
        _write_list_sheet(workbook.create_sheet("Errores"), document.errors)

    used_names = {sheet.title for sheet in workbook.worksheets}
    for table in document.tables:
        sheet_name = _unique_sheet_name(_safe_sheet_name(table.title), used_names)
        used_names.add(sheet_name)
        sheet = workbook.create_sheet(sheet_name)
        _write_table_sheet(sheet, table.columns, table.rows)

    for sheet in workbook.worksheets:
        _finish_sheet(sheet)

    workbook.save(output_path)

    return RenderResult(ok=True, format="xlsx", output_path=output_path, message="XLSX report generated.")


def _write_summary_sheet(sheet: Any, document: ReportDocument) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:D1")
    sheet["A1"] = document.title
    sheet["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor=DARK_FILL)
    sheet["A1"].alignment = Alignment(horizontal="left")

    sheet.merge_cells("A2:D2")
    sheet["A2"] = document.subtitle
    sheet["A2"].font = Font(size=11, italic=True, color="374151")
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    row = 4
    _write_section_title(sheet, row, "Estado general")
    row += 1
    _write_key_value_rows(
        sheet,
        row,
        [
            ("Tipo", document.report_type),
            ("Estado", document.status.upper()),
            ("Herramienta", document.meta.get("tool_name")),
            ("Módulo", document.meta.get("module_name")),
            ("Equipo", document.meta.get("computer_name")),
            ("Usuario", document.meta.get("user_name")),
            ("Generado/actualizado", _generated_at(document)),
            ("Schema", document.meta.get("schema_version")),
        ],
    )
    row += 9

    if document.summary:
        _write_section_title(sheet, row, "Resumen")
        row += 1
        _write_key_value_rows(sheet, row, list(document.summary.items()))
        row += len(document.summary) + 2

    brief = document.report_brief
    recommendations = brief.get("recommendations") if isinstance(brief.get("recommendations"), list) else []
    findings = brief.get("findings") if isinstance(brief.get("findings"), list) else []
    technician_notes = brief.get("technician_notes") if isinstance(brief.get("technician_notes"), list) else []
    if findings or recommendations or technician_notes:
        _write_section_title(sheet, row, "Lectura rápida")
        row += 1
        row = _write_bullets(sheet, row, "Hallazgos", findings)
        row = _write_bullets(sheet, row, "Recomendaciones", recommendations)
        row = _write_bullets(sheet, row, "Notas técnicas", technician_notes)

    sheet.freeze_panes = "A4"
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 34
    sheet.column_dimensions["C"].width = 22
    sheet.column_dimensions["D"].width = 22


def _write_disk_smart_overview(workbook: Workbook, document: ReportDocument) -> None:
    data = document.data
    disks = data.get("disks", []) if isinstance(data.get("disks"), list) else []
    sheet = workbook.create_sheet("Vista discos", 1)
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:J1")
    sheet["A1"] = "Vista ejecutiva de discos"
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor=DARK_FILL)

    metrics = [
        ("Discos", document.summary.get("disk_count")),
        ("OK", document.summary.get("ok_count")),
        ("Advertencias", document.summary.get("warnings_count")),
        ("Críticos", document.summary.get("critical_count")),
        ("Temperatura máx.", _temp(document.summary.get("max_temperature_c"))),
    ]
    col = 1
    for label, value in metrics:
        sheet.cell(row=3, column=col, value=label)
        sheet.cell(row=4, column=col, value=render_value(value))
        _style_metric_cell(sheet.cell(row=3, column=col), is_label=True)
        _style_metric_cell(sheet.cell(row=4, column=col), is_label=False)
        col += 2

    headers = [
        "#",
        "Estado",
        "Modelo",
        "Dispositivo",
        "Familia",
        "Temperatura",
        "Vida restante",
        "Horas encendido",
        "SMART",
        "Deduplicado",
    ]
    row = 7
    for idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=idx, value=header)
        _style_header_cell(cell)

    for disk_index, disk in enumerate(disks, start=1):
        if not isinstance(disk, dict):
            continue
        life = disk.get("life") if isinstance(disk.get("life"), dict) else {}
        values = [
            disk_index,
            str(disk.get("evaluation_status") or "INFO").upper(),
            disk.get("model"),
            disk.get("smart_device"),
            disk.get("storage_family") or disk.get("detected_type"),
            _temp(disk.get("temperature_c")),
            _percent(life.get("remaining_percent")),
            disk.get("power_on_hours"),
            "PASS" if disk.get("smart_global_passed") is True else "FAIL" if disk.get("smart_global_passed") is False else "",
            "Sí" if disk.get("deduplicated") is True else "No" if disk.get("deduplicated") is False else "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row + disk_index, column=col_idx, value=render_value(value))
            _style_body_cell(cell)
        _style_status_row(sheet, row + disk_index, str(disk.get("evaluation_status") or "INFO"), max_col=len(headers))

    if disks:
        sheet.auto_filter.ref = f"A7:J{7 + len(disks)}"
    sheet.freeze_panes = "A8"
    widths = [8, 14, 34, 16, 16, 14, 16, 18, 12, 14]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width


def _write_list_sheet(sheet: Any, items: list[Any]) -> None:
    if items and all(isinstance(item, dict) for item in items):
        columns: list[str] = []
        for item in items:
            for key in item.keys():
                key_text = str(key)
                if key_text not in columns:
                    columns.append(key_text)
        _write_table_sheet(sheet, columns, items)
    else:
        sheet.append(["Valor"])
        for item in items:
            sheet.append([render_value(item)])
        _format_basic_table(sheet)


def _write_table_sheet(sheet: Any, columns: list[str], rows: list[dict[str, Any]]) -> None:
    sheet.append(columns)
    for row in rows:
        sheet.append([render_value(row.get(column, "")) for column in columns])
    _format_basic_table(sheet)


def _write_key_value_rows(sheet: Any, start_row: int, rows: list[tuple[str, Any]]) -> None:
    sheet.cell(row=start_row, column=1, value="Campo")
    sheet.cell(row=start_row, column=2, value="Valor")
    _style_header_cell(sheet.cell(row=start_row, column=1))
    _style_header_cell(sheet.cell(row=start_row, column=2))
    row_number = start_row + 1
    for key, value in rows:
        if value in (None, "", [], {}):
            continue
        key_cell = sheet.cell(row=row_number, column=1, value=str(key))
        value_cell = sheet.cell(row=row_number, column=2, value=render_value(value))
        _style_body_cell(key_cell)
        _style_body_cell(value_cell)
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")
        row_number += 1


def _write_bullets(sheet: Any, row: int, title: str, items: list[Any]) -> int:
    if not items:
        return row
    sheet.cell(row=row, column=1, value=title)
    sheet.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    for item in items:
        sheet.cell(row=row, column=1, value="•")
        sheet.cell(row=row, column=2, value=render_value(item))
        sheet.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    return row + 1


def _write_section_title(sheet: Any, row: int, title: str) -> None:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = sheet.cell(row=row, column=1, value=title)
    cell.fill = PatternFill("solid", fgColor=SECTION_FILL)
    cell.font = Font(bold=True, color="111827")
    cell.alignment = Alignment(horizontal="left")


def _format_basic_table(sheet: Any) -> None:
    if sheet.max_row < 1:
        return
    for cell in sheet[1]:
        _style_header_cell(cell)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            _style_body_cell(cell)
        status_value = str(row[1].value if len(row) > 1 else "")
        if status_value:
            _style_status_row(sheet, row[0].row, status_value, max_col=sheet.max_column)
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    sheet.freeze_panes = "A2"


def _finish_sheet(sheet: Any) -> None:
    sheet.sheet_view.showGridLines = False
    thin = Side(style="thin", color=BORDER_COLOR)
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if cell.alignment is None or cell.alignment == Alignment():
                cell.alignment = Alignment(vertical="top")
    _autofit_columns(sheet)


def _autofit_columns(sheet: Any) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 10
        for cell in column_cells[:300]:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 42)


def _style_header_cell(cell: Any) -> None:
    cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_body_cell(cell: Any) -> None:
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def _style_metric_cell(cell: Any, *, is_label: bool) -> None:
    if is_label:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    else:
        cell.fill = PatternFill("solid", fgColor=INFO_FILL)
        cell.font = Font(size=14, bold=True, color="111827")
        cell.alignment = Alignment(horizontal="center")


def _style_status_row(sheet: Any, row: int, status: str, *, max_col: int) -> None:
    status_normalized = (status or "").strip().lower()
    fill = None
    if status_normalized in {"ok", "pass", "passed", "true"}:
        fill = PatternFill("solid", fgColor=OK_FILL)
    elif status_normalized in {"warning", "warn", "limited", "degraded"}:
        fill = PatternFill("solid", fgColor=WARNING_FILL)
    elif status_normalized in {"critical", "error", "fail", "failed"}:
        fill = PatternFill("solid", fgColor=CRITICAL_FILL)
    if fill:
        for col in range(1, max_col + 1):
            sheet.cell(row=row, column=col).fill = fill


def _safe_sheet_name(value: str) -> str:
    replacements = {
        "Atributos ATA": "Atributos ATA",
        "Métricas NVMe": "Metricas NVMe",
        "Dispositivos SMART alternativos": "SMART alternativos",
    }
    value = replacements.get(value, value)
    forbidden = set('[]:*?/\\')
    cleaned = "".join(" " if char in forbidden else char for char in value).strip()
    return (cleaned or "Datos")[:31]


def _unique_sheet_name(base_name: str, used_names: set[str]) -> str:
    if base_name not in used_names:
        return base_name
    index = 2
    while True:
        suffix = f" {index}"
        candidate = f"{base_name[:31 - len(suffix)]}{suffix}"
        if candidate not in used_names:
            return candidate
        index += 1


def _generated_at(document: ReportDocument) -> Any:
    return document.meta.get("generated_at") or document.meta.get("updated_at") or document.meta.get("created_at")


def _temp(value: Any) -> str | None:
    if value in (None, ""):
        return None
    return f"{value} °C"


def _percent(value: Any) -> str | None:
    if value in (None, ""):
        return None
    return f"{value}%"
